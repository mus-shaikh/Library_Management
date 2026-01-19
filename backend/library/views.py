from django.http import JsonResponse, HttpResponse
from .models import Author, Book, BorrowRecord
from openpyxl import Workbook


def get_authors(request):
    authors = list(Author.objects.values())
    return JsonResponse(authors, safe=False)


def get_books(request):
    books = list(Book.objects.values())
    return JsonResponse(books, safe=False)


def get_borrows(request):
    borrows = list(BorrowRecord.objects.values())
    return JsonResponse(borrows, safe=False)


def export_excel(request):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Authors"
    ws1.append(["ID", "Name", "Email", "Bio"])
    for a in Author.objects.all():
        ws1.append([a.id, a.name, a.email, a.bio])

    ws2 = wb.create_sheet("Books")
    ws2.append(["ID", "Title", "Genre", "Published", "Author"])
    for b in Book.objects.all():
        ws2.append([b.id, b.title, b.genre, b.published_date, b.author.name])

    ws3 = wb.create_sheet("BorrowRecords")
    ws3.append(["ID", "User", "Book", "Borrow Date", "Return Date"])
    for r in BorrowRecord.objects.all():
        ws3.append([r.id, r.user_name, r.book.title, r.borrow_date, r.return_date])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=library.xlsx"
    wb.save(response)
    return response
